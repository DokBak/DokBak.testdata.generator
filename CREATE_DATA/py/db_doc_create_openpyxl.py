import sys
import os
import logging
from datetime import datetime

# 現在のスクリプトのパスを基準にして、相対パスを取得
current_dir = os.path.dirname(os.path.abspath(__file__))

# ローカルパッケージのパスを追加
sys.path.append(os.path.join(current_dir, 'packages/openpyxl/'))
sys.path.append(os.path.join(current_dir, 'packages/et_xmlfile/'))

# ログファイルの設定
log_dir = os.path.join(current_dir, '../output/logs/')
os.makedirs(log_dir, exist_ok=True)  # ログディレクトリが存在しない場合、作成

# ログファイル名を設定
log_file_name = os.path.splitext(os.path.basename(__file__))[0] + ".log"
log_file_path = os.path.join(log_dir, log_file_name)

# 既存のログファイルをバックアップ
if os.path.exists(log_file_path):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_log_file = os.path.join(log_dir, f"bk_{timestamp}_{log_file_name}")
    os.rename(log_file_path, backup_log_file)  # 既存のログファイルをリネームしてバックアップ
    print(f"既存ログファイルをバックアップしました: {backup_log_file}")

# ログの設定
logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

logging.info("プログラムの実行開始")

# パッケージをインポート
import openpyxl

logging.info("Openpyxlパッケージのロード完了")

# Openpyxlバージョンを出力
print("Openpyxl バージョン:", openpyxl.__version__)
logging.info(f"Openpyxl バージョン: {openpyxl.__version__}")

# Excelファイルを作成
wb = openpyxl.Workbook()

# デフォルトシートを削除
wb.remove(wb.active)

# 新しいシートを作成
sheet_names = ["Version History", "Common Settings", "Table Name"]

for sheet_name in sheet_names:
    # 指定した名前のシートを作成
    ws = wb.create_sheet(title=sheet_name)
    # 各シートの全ての列幅を設定 (列幅 = 6)
    for col in range(1, 51):  # 1列目から51列目まで
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

    logging.info(f"シート '{sheet_name}' の作成完了")

# Excelファイルの保存先ディレクトリを設定
output_dir = os.path.join(current_dir, '../document/')
os.makedirs(output_dir, exist_ok=True)  # ディレクトリが存在しない場合、作成

# Excelファイルの保存パスを設定
output_file = os.path.join(output_dir, "sample.xlsx")

# 既存のExcelファイルをバックアップ
if os.path.exists(output_file):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_file = os.path.join(output_dir, f"bk_{timestamp}_sample.xlsx")
    os.rename(output_file, backup_file)
    logging.info(f"既存ファイルをバックアップしました: {backup_file}")

# 新しいExcelファイルを保存
wb.save(output_file)
print(f"Excelファイルが作成されました: {output_file}")
logging.info(f"Excelファイルが作成されました: {output_file}")

logging.info("プログラムの実行終了")
