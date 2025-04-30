import sys
import os
import logging
from datetime import datetime

# 現在のスクリプトのパスを基準にして、Excelファイルのパスを取得
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(current_dir, "../document/sample.xlsx")

# ローカルパッケージのパスを追加
sys.path.append(os.path.join(current_dir, 'packages/openpyxl/'))
sys.path.append(os.path.join(current_dir, 'packages/et_xmlfile/'))

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

# ログファイルの設定
log_dir = os.path.join(current_dir, "../output/logs/")
os.makedirs(log_dir, exist_ok=True)  # ログディレクトリを作成

# db_doc_set_common_openpyxl.py専用のログファイル名を設定
log_file_name = "db_doc_set_common_openpyxl.log"
log_file_path = os.path.join(log_dir, log_file_name)

# 既存のログファイルをバックアップ
if os.path.exists(log_file_path):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_log_file = os.path.join(log_dir, f"bk_{timestamp}_{log_file_name}")
    os.rename(log_file_path, backup_log_file)  # 既存のログファイルをリネームしてバックアップ

# ログ設定
logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# 結合セルの罫線情報
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ログ記録開始
logging.info("db_doc_set_common_openpyxl.py 実行開始")

try:
    # Excelファイルが存在するか確認
    if not os.path.exists(excel_file_path):
        error_message = f"Excelファイルが見つかりません: {excel_file_path}"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Excelファイルを開く
    logging.info(f"Excelファイルを開いています: {excel_file_path}")
    wb = openpyxl.load_workbook(excel_file_path)

    # Common Settings シートが存在するか確認
    if "Common Settings" not in wb.sheetnames:
        error_message = "Common Settings シートが見つかりません"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Common Settings シートを取得
    ws = wb["Common Settings"]

    # B2セルに '設定項目' を入力
    ws["B2"] = "設定項目"
    logging.info("B2 セルに '設定項目' を入力しました")

    # B2 から F2 をセル結合
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B2:F2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B2:F2 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B2:F2 にテーブル枠線と背景色を設定しました")

    # G2セルに '設定値' を入力
    ws["G2"] = "設定値"
    logging.info("G2 セルに '設定値' を入力しました")

    # G2 から M2 をセル結合
    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G2:M2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G2:M2 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G2:M2 にテーブル枠線と背景色を設定しました")

    # B3セルに 'データベース種類' を入力
    ws["B3"] = "データベース種類"
    logging.info("B3 セルに 'データベース種類' を入力しました")

    # B3 から F3 をセル結合
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B3:F3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B3:F3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B3:F3 にテーブル枠線と背景色を設定しました")

    # G3セルにデータベース種類を設定
    dv = DataValidation(type="list", formula1='"MySQL,MariaDB,PostgreSQL,SQLite,Oracle,Microsoft_SQL_Server,MongoDB,DynamoDB"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G3"])
    ws["G3"]="※データベース種類を選択"
    logging.info("G3 セルに 'データベース種類' を入力しました")

    # G3 から M3 をセル結合
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G3:M3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G3:M3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G3:M3 にテーブル枠線と背景色を設定しました")

    # B4セルに 'データベース名' を入力
    ws["B4"] = "データベース名"
    logging.info("B4 セルに 'データベース名' を入力しました")

    # B4 から F4 をセル結合
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B4:F4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B4:F4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B4:F4 にテーブル枠線と背景色を設定しました")

    # G4セルにデータベース名を設定
    ws["G4"] = "データベース名を入力"
    logging.info("G4 セルに 'データベース名' を入力しました")

    # G4 から M4 をセル結合
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G4:M4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G4:M4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G4:M4 にテーブル枠線と背景色を設定しました")

    # B5セルに 'スキーマ名' を入力
    ws["B5"] = "スキーマ名"
    logging.info("B5 セルに 'スキーマ名' を入力しました")

    # B5 から F5 をセル結合
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B5:F5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B5:F5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B5:F5 にテーブル枠線と背景色を設定しました")

    # G5セルにスキーマ名を設定
    ws["G5"] = "スキーマ名を入力"
    logging.info("G5 セルに 'スキーマ名' を入力しました")

    # G5 から M5 をセル結合
    ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G5:M5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G5:M5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G5:M5 にテーブル枠線と背景色を設定しました")

    # B6セルに 'データベースの文字コード' を入力
    ws["B6"] = "データベースの文字コード"
    logging.info("B6 セルに 'データベースの文字コード' を入力しました")

    # B6 から F6 をセル結合
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B6:F6 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B6"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B6:F6 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=6, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B6:F6 にテーブル枠線と背景色を設定しました")

    # G6セルにデータベースの文字コードを設定
    dv = DataValidation(type="list", formula1='"utf8mb4,utf8,sjis,euckr,UTF8,SJIS"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G6"])
    ws["G6"] = "※データベースの文字コードを選択"
    logging.info("G6 セルに 'データベースの文字コード' を入力しました")

    # G6 から M6 をセル結合
    ws.merge_cells(start_row=6, start_column=7, end_row=6, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G6:M6 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G6"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G6:M6 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=6, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G6:M6 にテーブル枠線と背景色を設定しました")

    # B7セルに 'データベースの並び順と比較方法' を入力
    ws["B7"] = "データベースの並び順と比較方法"
    logging.info("B7 セルに 'データベースの並び順と比較方法' を入力しました")

    # B7 から F7 をセル結合
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B7:F7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B7"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B7:F7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=7, max_row=7, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B7:F7 にテーブル枠線と背景色を設定しました")

    # G7セルにデータベースの並び順と比較方法を設定
    dv = DataValidation(type="list", formula1='"utf8mb4_unicode_ci,utf8mb4_general_ci,utf8mb4_bin,C,en_US.UTF-8,ko_KR.UTF-8,ja_JP.UTF-8"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G7"])
    ws["G7"] = "※データベースの並び順と比較方法を選択"
    logging.info("G7 セルに 'データベースの並び順と比較方法' を入力しました")

    # G7 から M7 をセル結合
    ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G7:M7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G7"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G7:M7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=7, max_row=7, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G7:M7 にテーブル枠線と背景色を設定しました")

    # B8セルに 'データベースの接続ユーザ' を入力
    ws["B8"] = "データベースの接続ユーザ"
    logging.info("B8 セルに 'データベースの接続ユーザ' を入力しました")

    # B8 から F8 をセル結合
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B8:F8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B8"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B8:F8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B8:F8 にテーブル枠線と背景色を設定しました")

    # G8セルにデータベースの接続ユーザを設定
    ws["G8"] = "データベースの接続ユーザを入力"
    logging.info("G8 セルに 'データベースの接続ユーザ' を入力しました")

    # G8 から M8 をセル結合
    ws.merge_cells(start_row=8, start_column=7, end_row=8, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G8:M8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G8"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G8:M8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G8:M8 にテーブル枠線と背景色を設定しました")

    # B9セルに 'データベースの接続パスワード' を入力
    ws["B9"] = "データベースの接続パスワード"
    logging.info("B9 セルに 'データベースの接続パスワード' を入力しました")

    # B9 から F9 をセル結合
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B9:F9 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B9"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B9:F9 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B9:F9 にテーブル枠線と背景色を設定しました")

    # G9セルにデータベースの接続パスワードを設定
    ws["G9"] = "データベースの接続パスワードを入力"
    logging.info("G9 セルに 'データベースの接続パスワード' を入力しました")

    # G9 から M9 をセル結合
    ws.merge_cells(start_row=9, start_column=7, end_row=9, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G9:M9 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G9"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G9:M9 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G9:M9 にテーブル枠線と背景色を設定しました")

    # B10セルに 'データベースのユーザの権限' を入力
    ws["B10"] = "データベースのユーザの権限"
    logging.info("B10 セルに 'データベースのユーザの権限' を入力しました")

    # B10 から F10 をセル結合
    ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B10:F10 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B10"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B10:F10 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=10, max_row=10, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B10:F10 にテーブル枠線と背景色を設定しました")

    # G10セルにデータベースのユーザの権限を設定
    dv = DataValidation(type="list", formula1='"ALL PRIVILEGES,SELECT,INSERT,UPDATE,DELETE,CREATE,DROP,Superuser,Createrole,CreateDB,Replication,BypassRLS"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G10"])
    ws["G10"] = "※データベースのユーザの権限を選択"
    logging.info("G10 セルに 'データベースのユーザの権限' を入力しました")

    # G10 から M10 をセル結合
    ws.merge_cells(start_row=10, start_column=7, end_row=10, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G10:M10 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G10"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G10:M10 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=10, max_row=10, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G10:M10 にテーブル枠線と背景色を設定しました")

    # B11セルに 'データベースのユーザのテーブル' を入力
    ws["B11"] = "データベースのユーザのテーブル"
    logging.info("B11 セルに 'データベースのユーザのテーブル' を入力しました")

    # B11 から F11 をセル結合
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B11:F11 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B11"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B11:F11 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=11, max_row=11, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B11:F11 にテーブル枠線と背景色を設定しました")

    # G11セルにデータベースのユーザのテーブルを設定
    dv = DataValidation(type="list", formula1='"*,特定テーブル名作成"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G11"])
    ws["G11"] = "※データベースのユーザのテーブルを選択"
    logging.info("G11 セルに 'データベースのユーザのテーブル' を入力しました")

    # G11 から M11 をセル結合
    ws.merge_cells(start_row=11, start_column=7, end_row=11, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G11:M11 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G11"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G11:M11 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=11, max_row=11, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G11:M11 にテーブル枠線と背景色を設定しました")

    # B12セルに '参照テーブル名' を入力
    ws["B12"] = "参照テーブル名"
    logging.info("B12 セルに '参照テーブル名' を入力しました")

    # B12 から F12 をセル結合
    ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B12:F12 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B12"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B12:F12 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=12, max_row=12, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B12:F12 にテーブル枠線と背景色を設定しました")

    # G12セルに参照テーブル名を設定
    ws["G12"] = "参照テーブル名を入力"
    logging.info("G12 セルに '参照テーブル名' を入力しました")

    # G12 から M12 をセル結合
    ws.merge_cells(start_row=12, start_column=7, end_row=12, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G12:M12 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G12"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G12:M12 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=12, max_row=12, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G12:M12 にテーブル枠線と背景色を設定しました")

    # B13セルに 'ユーザ接続ホスト' を入力
    ws["B13"] = "ユーザ接続ホスト"
    logging.info("B13 セルに 'ユーザ接続ホスト' を入力しました")

    # B13 から F13 をセル結合
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B13:F13 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B13"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B13:F13 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=13, max_row=13, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B13:F13 にテーブル枠線と背景色を設定しました")

    # G13セルにユーザ接続ホストを設定
    dv = DataValidation(type="list", formula1='"localhost,127.0.0.1,%,指定ホスト"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G13"])
    ws["G13"] = "※ユーザ接続ホストを選択"
    logging.info("G13 セルに 'ユーザ接続ホスト' を入力しました")

    # G13 から M13 をセル結合
    ws.merge_cells(start_row=13, start_column=7, end_row=13, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G13:M13 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G13"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G13:M13 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=13, max_row=13, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G13:M13 にテーブル枠線と背景色を設定しました")

    # B14セルに 'PostgreSQLテンプレート' を入力
    ws["B14"] = "PostgreSQLテンプレート"
    logging.info("B14 セルに 'PostgreSQLテンプレート' を入力しました")

    # B14 から F14 をセル結合
    ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B14:F14 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B14"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B14:F14 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=14, max_row=14, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B14:F14 にテーブル枠線と背景色を設定しました")

    # G14セルにPostgreSQLテンプレートを設定
    dv = DataValidation(type="list", formula1='"template0,template1"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G14"])
    ws["G14"] = "※PostgreSQLテンプレートを選択"
    logging.info("G14 セルに 'PostgreSQLテンプレート' を入力しました")

    # G14 から M14 をセル結合
    ws.merge_cells(start_row=14, start_column=7, end_row=14, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G14:M14 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G14"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G14:M14 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=14, max_row=14, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G14:M14 にテーブル枠線と背景色を設定しました")

    # B15セルに 'PostgreSQLコネクション' を入力
    ws["B15"] = "PostgreSQLコネクション"
    logging.info("B15 セルに 'PostgreSQLコネクション' を入力しました")

    # B15 から F15 をセル結合
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B15:F15 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B15"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B15:F15 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=15, max_row=15, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B15:F15 にテーブル枠線と背景色を設定しました")

    # G15セルにPostgreSQLコネクションを設定
    dv = DataValidation(type="list", formula1='"-1,接続数を入力"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G15"])
    ws["G15"] = "※PostgreSQLコネクションを選択"
    logging.info("G15 セルに 'PostgreSQLコネクション' を入力しました")

    # G15 から M15 をセル結合
    ws.merge_cells(start_row=15, start_column=7, end_row=15, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G15:M15 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G15"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G15:M15 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=15, max_row=15, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G15:M15 にテーブル枠線と背景色を設定しました")

    # B16セルに 'PostgreSQLコネクション許可' を入力
    ws["B16"] = "PostgreSQLコネクション許可"
    logging.info("B16 セルに 'PostgreSQLコネクション許可' を入力しました")

    # B16 から F16 をセル結合
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B16:F16 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B16"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B16:F16 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=16, max_row=16, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B16:F16 にテーブル枠線と背景色を設定しました")

    # G16セルにPostgreSQLコネクション許可を設定
    dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G16"])
    ws["G16"] = "※PostgreSQLコネクション許可を選択"
    logging.info("G16 セルに 'PostgreSQLコネクション許可' を入力しました")

    # G16 から M16 をセル結合
    ws.merge_cells(start_row=16, start_column=7, end_row=16, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G16:M16 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G16"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G16:M16 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=16, max_row=16, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G16:M16 にテーブル枠線と背景色を設定しました")

    # B17セルに '文字列タイプ' を入力
    ws["B17"] = "文字列タイプ"
    logging.info("B17 セルに '文字列タイプ' を入力しました")

    # B17 から F17 をセル結合
    ws.merge_cells(start_row=17, start_column=2, end_row=17, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B17:F17 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B17"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B17:F17 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=17, max_row=17, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B17:F17 にテーブル枠線と背景色を設定しました")

    # G17セルに文字列タイプを設定
    dv = DataValidation(type="list", formula1='"UPPER,LOWER,混在タイプはスペース入力"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G17"])
    ws["G17"] = "※文字列タイプを選択"
    logging.info("G17 セルに '文字列タイプ' を入力しました")

    # G17 から M17 をセル結合
    ws.merge_cells(start_row=17, start_column=7, end_row=17, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G17:M17 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G17"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G17:M17 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=17, max_row=17, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G17:M17 にテーブル枠線と背景色を設定しました")

    # B18セルに '出力データ件数' を入力
    ws["B18"] = "出力データ件数"
    logging.info("B18 セルに '出力データ件数' を入力しました")

    # B18 から F18 をセル結合
    ws.merge_cells(start_row=18, start_column=2, end_row=18, end_column=6)
    logging.info(f"シート 'Common Settings' のセル B18:F18 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B18"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル B18:F18 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=18, max_row=18, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Common Settings' のセル B18:F18 にテーブル枠線と背景色を設定しました")

    # G18セルに出力データ件数を設定
    ws["G18"] = "出力データ件数を入力"
    logging.info("G18 セルに '出力データ件数' を入力しました")

    # G18 から M18 をセル結合
    ws.merge_cells(start_row=18, start_column=7, end_row=18, end_column=13)
    logging.info(f"シート 'Common Settings' のセル G18:M18 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G18"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Common Settings' のセル G18:M18 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=18, max_row=18, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Common Settings' のセル G18:M18 にテーブル枠線と背景色を設定しました")

    # Excelファイルを保存
    wb.save(excel_file_path)
    logging.info(f"Excelファイルを更新しました: {excel_file_path}")
    print(f"Excelファイルを更新しました: {excel_file_path}")

except Exception as e:
    error_message = f"エラーが発生しました: {e}"
    print(error_message)
    logging.error(error_message)

# ログ記録終了
logging.info("db_doc_set_common_openpyxl.py 実行終了")
