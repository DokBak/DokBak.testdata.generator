import sys
import os
import logging
from datetime import datetime

# 現在のスクリプトのパスを基準にして、Excelファイルのパスを取得
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(current_dir, "../document/sample.xlsx")

# ローカルパッケージのパスを追加
sys.path.append(os.path.join(current_dir, 'packages/openpyxl/'))
sys.path.append(os.path.join(current_dir, 'packages/et_xmlfile/'))

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

# ログファイルの設定
log_dir = os.path.join(current_dir, "../output/log/")
os.makedirs(log_dir, exist_ok=True)  # ログディレクトリを作成

# db_doc_set_history_openpyxl.py専用のログファイル名を設定
log_file_name = "db_doc_set_history_openpyxl.log"
log_file_path = os.path.join(log_dir, log_file_name)

# 既存のログファイルをバックアップ
if os.path.exists(log_file_path):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_log_file = os.path.join(log_dir, f"bk_{timestamp}_{log_file_name}")
    os.rename(log_file_path, backup_log_file)  # 既存のログファイルをリネームしてバックアップ

# ログ設定
logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# ログ記録開始
logging.info("db_doc_set_history_openpyxl.py 実行開始")

try:
    # Excelファイルが存在するか確認
    if not os.path.exists(excel_file_path):
        error_message = f"Excelファイルが見つかりません: {excel_file_path}"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Excelファイルを開く
    logging.info(f"Excelファイルを開いています: {excel_file_path}")
    wb = openpyxl.load_workbook(excel_file_path)

    # Version History シートが存在するか確認
    if "Version History" not in wb.sheetnames:
        error_message = "Version History シートが見つかりません"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Version History シートを取得
    ws = wb["Version History"]

    # B2セルに 'FileName' を入力
    ws["B2"] = "ファイル名"
    logging.info("B2 セルに 'ファイル名' を入力しました")

    # B2 から V3 をセル結合
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=22)
    logging.info(f"シート 'Version History' のセル B2:V3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B2"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル B2:V3 を中央揃えに設定しました")

    # 結合セルの罫線情報
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=22):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル B2:V3 にテーブル枠線と背景色を設定しました")

    # W2セルに 'Version' を入力
    ws["W2"] = "バージョン"
    logging.info("W2 セルに 'バージョン' を入力しました")

    # W2 から AA3 をセル結合
    ws.merge_cells(start_row=2, start_column=23, end_row=3, end_column=27)
    logging.info(f"シート 'Version History' のセル W2:AA3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["W2"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル W2:AA3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=23, max_col=27):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル W2:AA3 にテーブル枠線と背景色を設定しました")

    # AB2セルに 'ModifiedDate' を入力
    ws["AB2"] = "修正日"
    logging.info("AB2 セルに '修正日' を入力しました")

    # AB2 から AF3 をセル結合
    ws.merge_cells(start_row=2, start_column=28, end_row=3, end_column=32)
    logging.info(f"シート 'Version History' のセル AB2:AF3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AB2"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AB2:AF3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=28, max_col=32):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル AB2:AF3 にテーブル枠線と背景色を設定しました")

    # AG2セルに 'CreateUser' を入力
    ws["AG2"] = "作成者"
    logging.info("AG2 セルに '作成者' を入力しました")

    # AG2 から AK3 をセル結合
    ws.merge_cells(start_row=2, start_column=33, end_row=3, end_column=37)
    logging.info(f"シート 'Version History' のセル AG2:AK3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AG2"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AG2:AK3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=33, max_col=37):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル AG2:AK3 にテーブル枠線と背景色を設定しました")

    # AL2セルに 'ModifiedUser' を入力
    ws["AL2"] = "修正者"
    logging.info("AL2 セルに '修正者' を入力しました")

    # AL2 から AP3 をセル結合
    ws.merge_cells(start_row=2, start_column=38, end_row=3, end_column=42)
    logging.info(f"シート 'Version History' のセル AL2:AP3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AL2"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AL2:AP3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=38, max_col=42):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル AL2:AP3 にテーブル枠線と背景色を設定しました")

    # B4セルに 'formula' を入力
    formula = '=MID(CELL("filename", $A$1), FIND("[", CELL("filename", $A$1)) + 1, FIND("]", CELL("filename", $A$1)) - FIND("[", CELL("filename", $A$1)) - 1)'
    ws["B4"] = formula
    logging.info("B4 セルに 'ファイル名' を入力しました")

    # B4 から V4 をセル結合
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=22)
    logging.info(f"シート 'Version History' のセル B4:V4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B4"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル B4:V4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=22):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル B4:V4 にテーブル枠線と背景色を設定しました")

    # W4セルに 'formula' を入力
    formula = '=LOOKUP(1,0/($D:$D<>""),$D:$D)'
    ws["W4"] = formula
    logging.info("W4 セルに 'バージョン' を入力しました")

    # W4 から AA4 をセル結合
    ws.merge_cells(start_row=4, start_column=23, end_row=4, end_column=27)
    logging.info(f"シート 'Version History' のセル W4:AA4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["W4"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル W4:AA4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=23, max_col=27):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル W4:AA4 にテーブル枠線と背景色を設定しました")

    # W4セルの形式
    ws["W4"].number_format = "0.0"

    # AB4セルに 'formula' を入力
    formula = '=LOOKUP(1,0/($F:$F<>""),$F:$F)'
    ws["AB4"] = formula
    logging.info("AB4 セルに '修正日' を入力しました")

    # AB4 から AF4 をセル結合
    ws.merge_cells(start_row=4, start_column=28, end_row=4, end_column=32)
    logging.info(f"シート 'Version History' のセル AB4:AF4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AB4"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AB4:AF4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=28, max_col=32):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル AB4:AF4 にテーブル枠線と背景色を設定しました")

    # AB4セルの形式
    ws["AB4"].number_format = "YYYY/MM/DD" 

    # AG4セルに 'formula' を入力
    formula = '=IF($AL$8="","",$AL$8)'
    ws["AG4"] = formula
    logging.info("AG4 セルに '作成者' を入力しました")

    # AG4 から AK4 をセル結合
    ws.merge_cells(start_row=4, start_column=33, end_row=4, end_column=37)
    logging.info(f"シート 'Version History' のセル AG4:AK4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AG4"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AG4:AK4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=33, max_col=37):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル AG4:AK4 にテーブル枠線と背景色を設定しました")

    # AL4セルに 'formula' を入力
    formula = '=LOOKUP(1,0/($AL:$AL<>""),$AL:$AL)'
    ws["AL4"] = formula
    logging.info("AL4 セルに '修正者' を入力しました")

    # AL4 から AP4 をセル結合
    ws.merge_cells(start_row=4, start_column=38, end_row=4, end_column=42)
    logging.info(f"シート 'Version History' のセル AL4:AP4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AL4"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AL4:AP4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=38, max_col=42):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル AL4:AP4 にテーブル枠線と背景色を設定しました")

    # B6セルに 'No.' を入力
    ws["B6"] = "No."
    logging.info("B6 セルに 'No.' を入力しました")

    # B6 から C7 をセル結合
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=3)
    logging.info(f"シート 'Version History' のセル B6:C7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル B6:C7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル B6:C7 にテーブル枠線と背景色を設定しました")

    # D6セルに 'Version' を入力
    ws["D6"] = "バージョン"
    logging.info("D6 セルに 'バージョン' を入力しました")

    # D6 から E7 をセル結合
    ws.merge_cells(start_row=6, start_column=4, end_row=7, end_column=5)
    logging.info(f"シート 'Version History' のセル D6:E7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["D6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル D6:E7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=4, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル D6:E7 にテーブル枠線と背景色を設定しました")

    # F6セルに 'ModifiedDate' を入力
    ws["F6"] = "修正日"
    logging.info("F6 セルに '修正日' を入力しました")

    # F6 から I7 をセル結合
    ws.merge_cells(start_row=6, start_column=6, end_row=7, end_column=9)
    logging.info(f"シート 'Version History' のセル F6:I7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["F6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル F6:I7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=6, max_col=9):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル F6:I7 にテーブル枠線と背景色を設定しました")

    # J6セルに 'ModifiedReason' を入力
    ws["J6"] = "修正理由"
    logging.info("J6 セルに '修正理由' を入力しました")

    # J6 から M7 をセル結合
    ws.merge_cells(start_row=6, start_column=10, end_row=7, end_column=13)
    logging.info(f"シート 'Version History' のセル J6:M7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["J6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル J6:M7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=10, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル J6:M7 にテーブル枠線と背景色を設定しました")

    # N6セルに 'ModifiedArea' を入力
    ws["N6"] = "修正シート"
    logging.info("N6 セルに '修正シート' を入力しました")

    # N6 から S7 をセル結合
    ws.merge_cells(start_row=6, start_column=14, end_row=7, end_column=19)
    logging.info(f"シート 'Version History' のセル N6:S7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル N6:S7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=14, max_col=19):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル N6:S7 にテーブル枠線と背景色を設定しました")

    # T6セルに 'ModifiedContents' を入力
    ws["T6"] = "修正内容"
    logging.info("T6 セルに '修正内容' を入力しました")

    # T6 から AK7 をセル結合
    ws.merge_cells(start_row=6, start_column=20, end_row=7, end_column=37)
    logging.info(f"シート 'Version History' のセル T6:AK7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["T6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル T6:AK7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=20, max_col=37):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル T6:AK7 にテーブル枠線と背景色を設定しました")

    # AL6セルに 'ModifiedUser' を入力
    ws["AL6"] = "修正者"
    logging.info("AL6 セルに '修正者' を入力しました")

    # AL6 から AP7 をセル結合
    ws.merge_cells(start_row=6, start_column=38, end_row=7, end_column=42)
    logging.info(f"シート 'Version History' のセル AL6:AP7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AL6"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AL6:AP7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=38, max_col=42):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Version History' のセル AL6:AP7 にテーブル枠線と背景色を設定しました")

    # B8セルに 'formula' を入力
    formula = '=ROW()-7'
    ws["B8"] = formula
    logging.info("B8 セルに 'No.' を入力しました")

    # B8 から C8 をセル結合
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
    logging.info(f"シート 'Version History' のセル B8:C8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル B8:C8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=2, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル B8:C8 にテーブル枠線と背景色を設定しました")

    # D8セルに 'formula' を入力
    formula = '=IF($B8<>"",IF($B8=1,1,$D7+0.1),)'
    ws["D8"] = formula
    logging.info("D8 セルに 'バージョン' を入力しました")

    # D8 から E8 をセル結合
    ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=5)
    logging.info(f"シート 'Version History' のセル D8:E8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["D8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル D8:E8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=4, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="868E96", end_color="868E96", fill_type="solid")
    logging.info("シート 'Version History' のセル D8:E8 にテーブル枠線と背景色を設定しました")

    # D8セルの形式
    ws["D8"].number_format = "0.0"

    # F8セルに実行日を設定
    ws["F8"] = datetime.now()
    logging.info("F8 セルに '実行日' を入力しました")

    # F8 から I8 をセル結合
    ws.merge_cells(start_row=8, start_column=6, end_row=8, end_column=9)
    logging.info(f"シート 'Version History' のセル F8:I8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["F8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル F8:I8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=6, max_col=9):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Version History' のセル F8:I8 にテーブル枠線と背景色を設定しました")

    # F8セルの形式
    ws["F8"].number_format = "YYYY/MM/DD" 

    # J8セルに修正理由を設定
    dv = DataValidation(type="list", formula1='"作成,修正,削除"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["J8"])
    logging.info("J8 セルに '作成理由' を入力しました")

    # J8 から M8 をセル結合
    ws.merge_cells(start_row=8, start_column=10, end_row=8, end_column=13)
    logging.info(f"シート 'Version History' のセル J8:M8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["J8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル J8:M8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=10, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Version History' のセル J8:M8 にテーブル枠線と背景色を設定しました")

    # N8セルに修正シートを設定
    ws["N8"] = ""
    logging.info("N8 セルに '修正シート' を入力しました")

    # N8 から S8 をセル結合
    ws.merge_cells(start_row=8, start_column=14, end_row=8, end_column=19)
    logging.info(f"シート 'Version History' のセル N8:S8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル N8:S8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=14, max_col=19):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Version History' のセル N8:S8 にテーブル枠線と背景色を設定しました")

    # T8セルに修正内容を設定
    ws["T8"] = ""
    logging.info("T8 セルに '修正内容' を入力しました")

    # T8 から AK8 をセル結合
    ws.merge_cells(start_row=8, start_column=20, end_row=8, end_column=37)
    logging.info(f"シート 'Version History' のセル T8:AK8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["T8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル T8:AK8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=20, max_col=37):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Version History' のセル T8:AK8 にテーブル枠線と背景色を設定しました")

    # AL8セルに修正者を設定
    ws["AL8"] = ""
    logging.info("AL8 セルに '修正者' を入力しました")

    # AL8 から AP8 をセル結合
    ws.merge_cells(start_row=8, start_column=38, end_row=8, end_column=42)
    logging.info(f"シート 'Version History' のセル AL8:AP8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["AL8"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.info(f"シート 'Version History' のセル AL8:AP8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=38, max_col=42):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Version History' のセル AL8:AP8 にテーブル枠線と背景色を設定しました")

    # Excelファイルを保存
    wb.save(excel_file_path)
    logging.info(f"Excelファイルを更新しました: {excel_file_path}")
    print(f"Excelファイルを更新しました: {excel_file_path}")

except Exception as e:
    error_message = f"エラーが発生しました: {e}"
    print(error_message)
    logging.error(error_message)

# ログ記録終了
logging.info("db_doc_set_history_openpyxl.py 実行終了")
