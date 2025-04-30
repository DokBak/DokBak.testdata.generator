import sys
import os
import logging
from datetime import datetime

# 現在のスクリプトのパスを基準にして、Excelファイルのパスを取得
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(current_dir, "../document/sample.xlsx")

# ローカルパッケージのパスを追加
sys.path.append(os.path.join(current_dir, 'packages/openpyxl/'))
sys.path.append(os.path.join(current_dir, 'packages/et_xmlfile/'))

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

# ログファイルの設定
log_dir = os.path.join(current_dir, "../output/logs/")
os.makedirs(log_dir, exist_ok=True)  # ログディレクトリを作成

# db_doc_set_table_openpyxl.py専用のログファイル名を設定
log_file_name = "db_doc_set_table_openpyxl.log"
log_file_path = os.path.join(log_dir, log_file_name)

# 既存のログファイルをバックアップ
if os.path.exists(log_file_path):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_log_file = os.path.join(log_dir, f"bk_{timestamp}_{log_file_name}")
    os.rename(log_file_path, backup_log_file)  # 既存のログファイルをリネームしてバックアップ

# ログ設定
logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# 結合セルの罫線情報
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ログ記録開始
logging.info("db_doc_set_table_openpyxl.py 実行開始")

try:
    # Excelファイルが存在するか確認
    if not os.path.exists(excel_file_path):
        error_message = f"Excelファイルが見つかりません: {excel_file_path}"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Excelファイルを開く
    logging.info(f"Excelファイルを開いています: {excel_file_path}")
    wb = openpyxl.load_workbook(excel_file_path)

    # Table Name シートが存在するか確認
    if "Table Name" not in wb.sheetnames:
        error_message = "Table Name シートが見つかりません"
        print(error_message)
        logging.error(error_message)
        exit(1)

    # Table Name シートを取得
    ws = wb["Table Name"]

    # B2セルに '設定項目' を入力
    ws["B2"] = "設定項目"
    logging.info("B2 セルに '設定項目' を入力しました")

    # B2 から F2 をセル結合
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
    logging.info(f"シート 'Table Name' のセル B2:F2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B2:F2 を中央揃えに設定しました")


    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B2:F2 にテーブル枠線と背景色を設定しました")

    # G2セルに '設定値' を入力
    ws["G2"] = "設定値"
    logging.info("G2 セルに '設定値' を入力しました")

    # G2 から M2 をセル結合
    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=13)
    logging.info(f"シート 'Table Name' のセル G2:M2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G2:M2 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル G2:M2 にテーブル枠線と背景色を設定しました")

    # B3セルに 'テーブル名(ファイル名)' を入力
    ws["B3"] = "テーブル名(ファイル名)"
    logging.info("B3 セルに 'テーブル名(ファイル名)' を入力しました")

    # B3 から F3 をセル結合
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    logging.info(f"シート 'Table Name' のセル B3:F3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B3:F3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B3:F3 にテーブル枠線と背景色を設定しました")

    # G3セルにテーブル名(ファイル名)を設定
    ws["G3"]="テーブル名(ファイル名)を入力"
    logging.info("G3 セルに 'テーブル名(ファイル名)' を入力しました")

    # G3 から M3 をセル結合
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=13)
    logging.info(f"シート 'Table Name' のセル G3:M3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G3:M3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G3:M3 にテーブル枠線と背景色を設定しました")

    # B4セルに 'ファイル拡張子' を入力
    ws["B4"] = "ファイル拡張子"
    logging.info("B4 セルに 'ファイル拡張子' を入力しました")

    # B4 から F4 をセル結合
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
    logging.info(f"シート 'Table Name' のセル B4:F4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B4:F4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B4:F4 にテーブル枠線と背景色を設定しました")

    # G4セルにファイル拡張子を設定
    dv = DataValidation(type="list", formula1='".csv,.tsv,.txt,拡張子なし"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G4"])
    ws["G4"]="※ファイル拡張子を選択"
    logging.info("G4 セルに 'ファイル拡張子' を入力しました")

    # G4 から M4 をセル結合
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=13)
    logging.info(f"シート 'Table Name' のセル G4:M4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G4:M4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G4:M4 にテーブル枠線と背景色を設定しました")

    # B5セルに '圧縮形式' を入力
    ws["B5"] = "圧縮形式"
    logging.info("B5 セルに '圧縮形式' を入力しました")

    # B5 から F5 をセル結合
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
    logging.info(f"シート 'Table Name' のセル B5:F5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B5:F5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B5:F5 にテーブル枠線と背景色を設定しました")

    # G5セルに圧縮形式を設定
    dv = DataValidation(type="list", formula1='".gz,.zip,.Z,.tar,.7z,拡張子なし"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G5"])
    ws["G5"]="※圧縮形式を選択"
    logging.info("G5 セルに '圧縮形式' を入力しました")

    # G5 から M5 をセル結合
    ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=13)
    logging.info(f"シート 'Table Name' のセル G5:M5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G5:M5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G5:M5 にテーブル枠線と背景色を設定しました")

    # B6セルに '改行コード' を入力
    ws["B6"] = "改行コード"
    logging.info("B6 セルに '改行コード' を入力しました")

    # B6 から F6 をセル結合
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=6)
    logging.info(f"シート 'Table Name' のセル B6:F6 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B6"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B6:F6 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=6, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B6:F6 にテーブル枠線と背景色を設定しました")

    # G6セルに改行コードを設定
    dv = DataValidation(type="list", formula1='"CRLF,LF"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G6"])
    ws["G6"]="※改行コードを選択"
    logging.info("G6 セルに '改行コード' を入力しました")

    # G6 から M6 をセル結合
    ws.merge_cells(start_row=6, start_column=7, end_row=6, end_column=13)
    logging.info(f"シート 'Table Name' のセル G6:M6 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G6"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G6:M6 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=6, max_row=6, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G6:M6 にテーブル枠線と背景色を設定しました")

    # B7セルに '文字コード' を入力
    ws["B7"] = "文字コード"
    logging.info("B7 セルに '文字コード' を入力しました")

    # B7 から F7 をセル結合
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=6)
    logging.info(f"シート 'Table Name' のセル B7:F7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B7"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B7:F7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=7, max_row=7, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B7:F7 にテーブル枠線と背景色を設定しました")

    # G7セルに文字コードを設定
    dv = DataValidation(type="list", formula1='"UTF-8,UTF-16,EUC-JP,SJIS,CP932,EUC-KR"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G7"])
    ws["G7"]="※文字コードを選択"
    logging.info("G7 セルに '文字コード' を入力しました")

    # G7 から M7 をセル結合
    ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=13)
    logging.info(f"シート 'Table Name' のセル G7:M7 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G7"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G7:M7 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=7, max_row=7, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G7:M7 にテーブル枠線と背景色を設定しました")

    # B8セルに '区切り文字' を入力
    ws["B8"] = "区切り文字"
    logging.info("B8 セルに '区切り文字' を入力しました")

    # B8 から F8 をセル結合
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=6)
    logging.info(f"シート 'Table Name' のセル B8:F8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B8"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B8:F8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B8:F8 にテーブル枠線と背景色を設定しました")

    # G8セルに区切り文字を設定
    dv = DataValidation(type="list", formula1='"COMMA,TAB,拡張子なし"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G8"])
    ws["G8"]="※区切り文字を選択"
    logging.info("G8 セルに '区切り文字' を入力しました")

    # G8 から M8 をセル結合
    ws.merge_cells(start_row=8, start_column=7, end_row=8, end_column=13)
    logging.info(f"シート 'Table Name' のセル G8:M8 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G8"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G8:M8 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G8:M8 にテーブル枠線と背景色を設定しました")

    # B9セルに '囲み文字' を入力
    ws["B9"] = "囲み文字"
    logging.info("B9 セルに '囲み文字' を入力しました")

    # B9 から F9 をセル結合
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=6)
    logging.info(f"シート 'Table Name' のセル B9:F9 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["B9"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル B9:F9 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=2, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル B9:F9 にテーブル枠線と背景色を設定しました")

    # G9セルに囲み文字を設定
    dv = DataValidation(type="list", formula1='"DOUBLE_QUOTE,SINGLE_QUOTE,拡張子なし"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["G9"])
    ws["G9"]="※囲み文字を選択"
    logging.info("G9 セルに '囲み文字' を入力しました")

    # G9 から M9 をセル結合
    ws.merge_cells(start_row=9, start_column=7, end_row=9, end_column=13)
    logging.info(f"シート 'Table Name' のセル G9:M9 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["G9"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル G9:M9 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=9, max_row=9, min_col=7, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル G9:M9 にテーブル枠線と背景色を設定しました")

    # N2セルに '設定項目' を入力
    ws["N2"] = "設定項目"
    logging.info("N2 セルに '設定項目' を入力しました")

    # N2 から R2 をセル結合
    ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=18)
    logging.info(f"シート 'Table Name' のセル N2:R2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル N2:R2 を中央揃えに設定しました")


    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=14, max_col=18):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル N2:R2 にテーブル枠線と背景色を設定しました")

    # S2セルに '設定値' を入力
    ws["S2"] = "設定値"
    logging.info("S2 セルに '設定値' を入力しました")

    # S2 から Y2 をセル結合
    ws.merge_cells(start_row=2, start_column=19, end_row=2, end_column=25)
    logging.info(f"シート 'Table Name' のセル S2:Y2 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["S2"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル S2:Y2 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=19, max_col=25):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル S2:Y2 にテーブル枠線と背景色を設定しました")

    # N3セルに '日付の区分文字' を入力
    ws["N3"] = "日付の区分文字"
    logging.info("N3 セルに '日付の区分文字' を入力しました")

    # N3 から R3 をセル結合
    ws.merge_cells(start_row=3, start_column=14, end_row=3, end_column=18)
    logging.info(f"シート 'Table Name' のセル N3:R3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル N3:R3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=14, max_col=18):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル N3:R3 にテーブル枠線と背景色を設定しました")

    # S3セルに日付の区分文字を設定
    dv = DataValidation(type="list", formula1='"/,-,SPACE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["S3"])
    ws["S3"]="日付の区分文字を選択"
    logging.info("S3 セルに '日付の区分文字' を入力しました")

    # S3 から Y3 をセル結合
    ws.merge_cells(start_row=3, start_column=19, end_row=3, end_column=25)
    logging.info(f"シート 'Table Name' のセル S3:Y3 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["S3"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル S3:Y3 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=19, max_col=25):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル S3:Y3 にテーブル枠線と背景色を設定しました")

    # N4セルに '時間の区分文字' を入力
    ws["N4"] = "時間の区分文字"
    logging.info("N4 セルに '時間の区分文字' を入力しました")

    # N4 から R4 をセル結合
    ws.merge_cells(start_row=4, start_column=14, end_row=4, end_column=18)
    logging.info(f"シート 'Table Name' のセル N4:R4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル N4:R4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=14, max_col=18):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル N4:R4 にテーブル枠線と背景色を設定しました")

    # S4セルに時間の区分文字を設定
    dv = DataValidation(type="list", formula1='":,SPACE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["S4"])
    ws["S4"]="時間の区分文字を選択"
    logging.info("S4 セルに '時間の区分文字' を入力しました")

    # S4 から Y4 をセル結合
    ws.merge_cells(start_row=4, start_column=19, end_row=4, end_column=25)
    logging.info(f"シート 'Table Name' のセル S4:Y4 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["S4"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル S4:Y4 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=19, max_col=25):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル S4:Y4 にテーブル枠線と背景色を設定しました")

    # N5セルに '日付と時間の間の文字' を入力
    ws["N5"] = "日付と時間の間の文字"
    logging.info("N5 セルに '日付と時間の間の文字' を入力しました")

    # N5 から R5 をセル結合
    ws.merge_cells(start_row=5, start_column=14, end_row=5, end_column=18)
    logging.info(f"シート 'Table Name' のセル N5:R5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["N5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル N5:R5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=14, max_col=18):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")
    logging.info("シート 'Table Name' のセル N5:R5 にテーブル枠線と背景色を設定しました")

    # S5セルに日付と時間の間の文字を設定
    dv = DataValidation(type="list", formula1='"拡張子なし,SPACE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["S5"])
    ws["S5"]="日付と時間の間の文字を選択"
    logging.info("S5 セルに '日付と時間の間の文字' を入力しました")

    # S5 から Y5 をセル結合
    ws.merge_cells(start_row=5, start_column=19, end_row=5, end_column=25)
    logging.info(f"シート 'Table Name' のセル S5:Y5 を結合しました")

    # 結合セルの中央揃え設定
    merged_cell = ws["S5"]
    merged_cell.alignment = Alignment(horizontal="left", vertical="center")
    logging.info(f"シート 'Table Name' のセル S5:Y5 を中央揃えに設定しました")

    # 結合セルの罫線と背景色設定
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=19, max_col=25):
        for cell in row:
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    logging.info("シート 'Table Name' のセル S5:Y5 にテーブル枠線と背景色を設定しました")

    # Excelファイルを保存
    wb.save(excel_file_path)
    logging.info(f"Excelファイルを更新しました: {excel_file_path}")
    print(f"Excelファイルを更新しました: {excel_file_path}")

except Exception as e:
    error_message = f"エラーが発生しました: {e}"
    print(error_message)
    logging.error(error_message)

# ログ記録終了
logging.info("db_doc_set_table_openpyxl.py 実行終了")
